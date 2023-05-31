#!/usr/bin/env python

import os
from openpyxl import Workbook
from openpyxl import load_workbook

SourceData = '.\\Lottery_Powerball_Winning_Numbers__Beginning_2010.csv'
OutputData = '.\\Lottery_Powerball_Modified.csv'

os.chdir(WorkingDirectory)

workbook = load_workbook(filename = WorkingDirectory + '\\DataSummary.xlsx')
workbook.sheetnames
#['SourceData']

sheet = workbook.active
print(sheet)

print(sheet.title)
